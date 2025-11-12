from openpyxl import load_workbook
from copy import copy
from pathlib import Path
from re import match
from os import listdir
import sys

"""
版本：     0.1
修订日期：  2025-6-11
记录：     功能实现
"""

# 获取自身路径
PATH = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent


def find_path(dir_path=PATH, key='', end='') -> Path:
    name = []
    for path_name in (name_list := listdir(dir_path)):
        if match(rf".*?{key}.*?\.{end}", path_name):
            name.append(path_name)
    if (length := len(name)) == 0:
        print(f"该目录文件有：{name_list}")
        if key == '':
            raise ValueError(f"无法匹配后缀为：'{end}'的文件")
        else:
            raise ValueError(f"无法匹配关键词为：'{key}'的文件")
    elif length == 1:
        return Path(dir_path / name[0])
    else:
        print('匹配到多个结果：')
        for _i in range(1, length + 1):
            print(f"{_i}、{name[_i - 1]}")
        in_num = int(input('选择：')) - 1
        if in_num < 0 or in_num >= length:
            raise ValueError('不存在该选项')
        return Path(dir_path / name[in_num])


def adjust_row_heights(file_path, output_path, scale_factor=1.1):
    """  调整Excel所有行高为原值的110%    """
    wb = load_workbook(filename=file_path)
    for sheet in wb._sheets[3:]:
        print(f"--> 正在处理工作表: {sheet.title} <--")
        for row in sheet.iter_rows(min_row=3):
            original_height = sheet.row_dimensions[row[0].row].height
            if original_height is not None:
                # 设置新行高（原值×缩放系数）
                new_height = original_height * scale_factor if original_height > 25 else 25
            else:
                new_height = 25
            sheet.row_dimensions[row[0].row].height = new_height
            # 可选：设置自动换行确保内容可见
            for cell in row:
                # cell.alignment = Alignment(wrap_text=True)
                cell.font = copy(cell.font)
                cell.border = copy(cell.border)
                cell.fill = copy(cell.fill)
                cell.number_format = cell.number_format
                cell.protection = copy(cell.protection)
                cell.alignment = copy(cell.alignment)
    # 保存结果
    wb.save(output_path)
    print(f"处理完成，结果已保存至: {output_path}")
    return


if __name__ == "__main__":
    input_file = find_path(key="^S", end="xlsx")
    output_file = PATH / (input_file.stem + "_adjusted.xlsx")
    adjust_row_heights(input_file, output_file)